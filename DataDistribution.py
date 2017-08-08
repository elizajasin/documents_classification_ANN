__author__ = 'elizajasin'

import os
import openpyxl
from openpyxl import Workbook

os.getcwd
wb = openpyxl.load_workbook('hadits_fix.xlsx')
ws = wb.active
informasi = []
anjuran = []
larangan = []
for i in range (2,ws.max_row+1):
    if ws.cell(row=i,column=ws.max_column).value == 1:
        informasi.append(str(ws.cell(row=i,column=ws.max_column-1).value))
    elif ws.cell(row=i,column=ws.max_column).value == 2:
        anjuran.append(str(ws.cell(row=i,column=ws.max_column-1).value))
    else:
        larangan.append(str(ws.cell(row=i,column=ws.max_column-1).value))

# wb = openpyxl.load_workbook('hadits_fix_7008_balance.xlsx')
# ws = wb.active
# for i in range (2,ws.max_row+1):
#     if ws.cell(row=i,column=ws.max_column).value == 1:
#         informasi.append(str(ws.cell(row=i,column=ws.max_column-1).value))
#     elif ws.cell(row=i,column=ws.max_column).value == 2:
#         anjuran.append(str(ws.cell(row=i,column=ws.max_column-1).value))
#     else:
#         larangan.append(str(ws.cell(row=i,column=ws.max_column-1).value))
#
# wb = openpyxl.load_workbook('hadits_fix_3.xlsx')
# ws = wb.active
# for i in range (2,ws.max_row+1):
#     if ws.cell(row=i,column=ws.max_column).value == 1:
#         informasi.append(str(ws.cell(row=i,column=ws.max_column-1).value))
#     elif ws.cell(row=i,column=ws.max_column).value == 2:
#         anjuran.append(str(ws.cell(row=i,column=ws.max_column-1).value))
#     else:
#         larangan.append(str(ws.cell(row=i,column=ws.max_column-1).value))

wb = Workbook()
ws = wb.active
for i in range(243):
    data = []
    data.append(informasi[i])
    data.append(1)
    ws.append(data)
    data = []
    data.append(anjuran[i])
    data.append(2)
    ws.append(data)
    data = []
    data.append(larangan[i])
    data.append(3)
    ws.append(data)
wb.save('hadits_merge_fix.xlsx')

print(len(informasi))
print(len(anjuran))
print(len(larangan))