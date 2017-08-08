__author__ = 'elizajasin'

import os
import openpyxl
import ReadAndWriteFile as RAWfile

def readDataClass (filename):
    os.getcwd
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    hadits = []
    for row in range(1, ws.max_row+1):
        cell_name = "{}{}".format('B', row)
        hadits.append(ws[cell_name].value)
    return hadits

kelas = readDataClass("hadits_merge_fix.xlsx")
biner_class = []
for i in range(len(kelas)):
    if kelas[i] == 1 :
        temp = [1,0,0]
    elif kelas[i] == 2:
        temp = [0,1,0]
    else:
        temp = [0,0,1]
    biner_class.append(temp)
RAWfile.writeData(biner_class,"data_target_merge_fix.xlsx")
print(len(biner_class))
print(biner_class)