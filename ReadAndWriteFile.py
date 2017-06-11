__author__ = 'elizajasin'

import os
import openpyxl
from  openpyxl import Workbook
import pandas

def readData (filename):
    os.getcwd
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    hadits = []
    for i in range (2,ws.max_row+1):
        hadits.append(str(ws.cell(row=i,column=ws.max_column-1).value))
    return hadits

def writeData (data, filename):
    wb = Workbook()
    ws = wb.active
    for i in range(len(data)):
        ws.append(data[i])
    wb.save(filename)

def readDataPreprocessing (filename):
    os.getcwd
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    hadits = []
    for i in range (1,ws.max_row+1):
        s = []
        for j in range(1,ws.max_column):
            s.append(str(ws.cell(row=i,column=j).value))
        hadits.append(s)
    for i in range(len(hadits)):
        while 'None' in hadits[i]:
            hadits[i].remove('None')
    return hadits

def writeFEResult (atribut, filename):
    wb = Workbook()
    ws = wb.active
    for key in atribut:
        ws.append(atribut[key])
    wb.save(filename)

def writeFEResultWithPandas (atribut, filename):
    ws = []
    for key in atribut:
        ws.append(atribut[key])
    df = pandas.DataFrame(ws)
    writer = pandas.ExcelWriter(filename)
    df.to_excel(writer,'Sheet1')
    writer.save()

def writeFEResultWithTxt (atribut, filename):
    file = open(filename, 'w')
    for key in atribut:
        for i in range(len(atribut[key])):
            file.write(str(atribut[key][i]) + ' ')
        file.write('\n')
    file.close()

def readDataClass (filename):
    os.getcwd
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    hadits = []
    for row in range(2, ws.max_row+1):
        cell_name = "{}{}".format('D', row)
        hadits.append(ws[cell_name].value)
    return hadits

def readDataInputTrain (filename):
    os.getcwd
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    hadits = []
    for i in range (1,ws.max_row+1):
        s = []
        for j in range(1, ws.max_column):
            s.append(str(ws.cell(row=i, column=j).value))
        hadits.append(s)
    return hadits