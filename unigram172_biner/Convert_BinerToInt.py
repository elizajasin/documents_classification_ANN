__author__ = 'elizajasin'

import os
import openpyxl
from openpyxl import Workbook

def openfile (filename):
    os.getcwd
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    biner = []
    for i in range(1, ws.max_row + 1):
        s = []
        for j in range(1, ws.max_column + 1):
            s.append(ws.cell(row=i, column=j).value)
        biner.append(s)
    return biner

def convert_to_int (biner):
    inte = []
    for i in range(len(biner)):
        if biner[i][0] == None:
            inte.append(0)
        else:
            if (biner[i][0] == 1) and (biner[i][1] == 0) and (biner[i][2] == 0):
                inte.append(1)
            elif (biner[i][0] == 0) and (biner[i][1] == 1) and (biner[i][2] == 0):
                inte.append(2)
            else:
                inte.append(3)
    return inte

binerTrain = openfile('trainTarget_unigram172.xlsx')
binerVal = openfile('valTarget_unigram172.xlsx')
binerTest = openfile('testTarget_unigram172.xlsx')

inteTrain = convert_to_int(binerTrain)
inteVal = convert_to_int(binerVal)
inteTest = convert_to_int(binerTest)

if len(inteTrain) != 516:
    for i in range(516-len(inteTrain)):
        inteTrain.append(0)
if len(inteVal) != 516:
    for i in range(516-len(inteVal)):
        inteVal.append(0)
if len(inteTest) != 516:
    for i in range(516-len(inteTest)):
        inteTest.append(0)

inte = []
for i in range(516):
    s = []
    s.append(inteTrain[i])
    s.append(inteVal[i])
    s.append(inteTest[i])
    inte.append(s)

wb = Workbook()
ws = wb.active
for i in range(len(inte)):
    ws.append(inte[i])
wb.save('result_integer.xlsx')