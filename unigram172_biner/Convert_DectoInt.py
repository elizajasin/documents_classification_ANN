__author__ = 'elizajasin'

import os
import numpy as np
import openpyxl
from openpyxl import Workbook

def openfile (filename):
    os.getcwd
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    dec = []
    for i in range(1, ws.max_row + 1):
        s = []
        for j in range(1, ws.max_column + 1):
            s.append(ws.cell(row=i, column=j).value)
        dec.append(s)
    return dec

def convert_to_int (dec):
    inte = []
    for i in range(len(dec)):
        a = max(dec[i])
        inte.append(dec[i].index(a)+1)
    return inte

dec = openfile('output_unigram172.xlsx')
inte = convert_to_int(dec)

wb = Workbook()
ws = wb.active
for i in range(len(inte)):
    ws.cell(row=i+1, column=1).value = inte[i]
wb.save('result_integer_ouput.xlsx')