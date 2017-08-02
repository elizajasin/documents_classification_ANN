__author__ = 'elizajasin'

from string import punctuation
import re
import os
import openpyxl
from openpyxl import Workbook
from py2casefold import casefold
import Preprocessing as PreP
import FeatureExtraction as FE

def readData (filename):
    os.getcwd
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    hadits = []
    for i in range (1,4):
        hadits.append(str(ws.cell(row=i,column=3).value))
    return hadits

def writeData (data, filename):
    wb = Workbook()
    ws = wb.active
    for i in range(len(data)):
        ws.append(data[i])
    wb.save(filename)

def removePunct (hadits):
    punct = list(punctuation)
    punct.remove("-")
    punct = ''.join(punct)
    remove_punct = re.compile('[%s]' % re.escape(punctuation)).sub('', hadits)
    remove_punct = re.split(r'\s', remove_punct)
    return remove_punct

data = readData('Book1.xlsx')
rempun = []
for i in range(len(data)):
    rempun.append(removePunct(data[i]))
# writeData(rempun,'Book1_removepunc.xlsx')
for i in range(len(rempun)):
    for j in range(len(rempun[i])):
        rempun[i][j] = casefold(rempun[i][j])
# writeData(rempun,'Book1_casefold.xlsx')
stopword_remove = PreP.stopwordsRemoval('id.stopwords.02.01.2016.txt',rempun)
# writeData(stopword_remove,'Book1_stopword.xlsx')
lemma = PreP.lemmatization(stopword_remove)
# writeData(lemma,'Book1_lemma.xlsx')
remove_kbbi = PreP.stopword_kbbi(lemma)
# writeData(remove_kbbi,'Book1_lemma.xlsx')
atribut1 = FE.makeAtribut(remove_kbbi)
atribut2 = FE.makeAtributBigram(remove_kbbi)
atribut3 = FE.makeAtributTrigram(remove_kbbi)
FE_result1 = FE.sumFE(remove_kbbi,atribut1)
FE_result2 = FE.sumFE(remove_kbbi,atribut2)
FE_result3 = FE.sumFE(remove_kbbi,atribut3)
print(FE_result1)
print(FE_result2)
print(FE_result3)